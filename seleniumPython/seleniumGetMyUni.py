"""
The below program gets data from excel sheet and the look up for the data on a website and then scraps that data
and then insert it on a excel sheet.
"""
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as op
from selenium import webdriver

drive = "" #! Please enter drive path.
fileName = "" #! Please enter file name.
path = drive + fileName

r = 3 #! Please change the row number to the row number from which you would like to start reading the file.
nameSet = set()
def institutionNames():
    """
    This function load all the data from a excel sheet for a particular column and stores it in a list named "nameList".
    Current design to get data from single column but can be chaged to get data from multiple column.
    """
    itn = op.load_workbook(path)
    itnsh = itn["Sheet1"]
    maxRow = 1500 # itnsh.max_row
    for s in range(r, (maxRow + 5)):
        name = itnsh.cell(row = s, column = 3).value #! Please change the column number based on the column number in your file.
        if name != None:
            name = name.strip()
            nameSet.add(name)
    itn.close()
    # print(nameSet)
    return nameSet

appByDict1 = dict()
affByDict1 = dict()
appByDict2 = dict()
affByDict2 = dict()
def dataScreapping():
    """
    This function will scrap the data from the website and store it a dict name "accreditationsDict".
    It uses selenium webdriver for searching and getting the data.
    """
    chromeDriverPath = "" #! Please add your googlechrome driver path.
    chromeOptions = webdriver.ChromeOptions()
    #? Feel free to adjust the below argument based on your requirement.
    chromeOptions.add_argument("start-maximized")
    chromeOptions.add_argument("disable-infobars")
    chromeOptions.add_argument("disable-notifications")
    chromeOptions.add_argument("disable-popup-blocking")
    # chromeOptions.add_argument("headless")
    chromeOptions.add_experimental_option("detach", True)

    driver = webdriver.Chrome(chromeDriverPath, options=chromeOptions)
    driver.get("https://www.getmyuni.com/")

    #? Getting the below data from the above function. So please call the above function before calling this function.
    nameSetLength = len(nameSet)
    print(nameSetLength)
    curr = 0
    for i in nameSet:
        curr += 1
        print(f"{curr} - {i}")
        driver.get("https://www.getmyuni.com/")

        # time.sleep(1)
        searchBarID = "coll-name-box"
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, searchBarID)))
        searchBar = driver.find_element_by_id(searchBarID)
        searchBar.send_keys(i)

        try:
            # time.sleep(5)
            #! In the below line change div[8] to div[7], if the function returns "Institution not found" consistantly.
            #! It may be because of div position has changed.
            #? Please choose one of the below versions of Xpath.
            searchedNameXpath = "/html/body/div[7]/section/div[2]/div[2]/div[4]/a" 
            # searchedNameXpath ="//*[@id='search-div']/section/div[2]/div[2]/div[4]/a[1]"
            WebDriverWait(driver, 8).until(EC.element_to_be_clickable((By.XPATH, searchedNameXpath)))
            searchedName = driver.find_element_by_xpath(searchedNameXpath)
            searchedName.click()

            time.sleep(2)
            #? Getting data from the first element.
            try:
                affEle1Xpath = "/html/body/div[5]/div[1]/div/div/div/div/div/div/div[2]/div[1]/div/strong[1]"
                affEle1 = driver.find_element_by_xpath(affEle1Xpath)
                affEleText1 = affEle1.text
                # print(affEleText1)
                
                if affEleText1 == "Approved By:":
                    appBy = driver.find_element_by_xpath("/html/body/div[5]/div[1]/div/div/div/div/div/div/div[2]/div[1]/div/span")
                    appByText = appBy.text
                    appByDict1[i] = appByText
                elif affEleText1 == "Affiliated By:":
                    affBy = driver.find_element_by_xpath("/html/body/div[5]/div[1]/div/div/div/div/div/div/div[2]/div[1]/div/a")
                    affByText = affBy.text
                    affByDict1[i] = affByText
                else:
                    print("First data unavailable.")
            except:
                pass

            #? Getting data from the second element.
            try:
                affEle2 = driver.find_element_by_xpath("/html/body/div[5]/div[1]/div/div/div/div/div/div/div[2]/div[1]/div/strong[2]")
                affEleText2 = affEle2.text

                if affEleText2 == "Approved By:":
                    appBy = driver.find_element_by_xpath("/html/body/div[5]/div[1]/div/div/div/div/div/div/div[2]/div[1]/div/span")
                    appByText = appBy.text
                    appByDict2[i] = appByText
                elif affEleText2 == "Affiliated By:":
                    affBy = driver.find_element_by_xpath("/html/body/div[5]/div[1]/div/div/div/div/div/div/div[2]/div[1]/div/a")
                    affByText = affBy.text
                    affByDict2[i] = affByText
                else:
                    pass
            except:
                print("Second data not available.")
        except:
            print("Institution not found!!!")
    return appByDict1, affByDict1, appByDict2, affByDict2,

def dataPosting():
    """
    This functions open the same excel sheet from which we retrieve data
    and then will store the collected data in the new column.
    """
    itn = op.load_workbook(path)
    itnSh = itn["Sheet1"]
    maxRow = itnSh.max_row
    for s in range(r, (maxRow + 10)):
        name = itnSh.cell(row = s, column = 3).value
        for key, val in appByDict1.items():
            if key == name:
                approvedBy1 = itnSh.cell(row = s, column = 11)
                approvedBy1.value = val
        for key, val in affByDict1.items():
            if key == name:
                affiliatedBy1 = itnSh.cell(row = s, column = 12)
                affiliatedBy1.value = val
        for key, val in appByDict2.items():
            if key == name:
                approvedBy2 = itnSh.cell(row = s, column = 11)
                approvedBy2.value = val
        for key, val in affByDict2.items():
            if key == name:
                affiliatedBy2 = itnSh.cell(row = s, column = 12)
                affiliatedBy2.value = val
    itn.save(path)
    itn.close()
    print("All Done!!!")

#? Calling the above functions, which will get data from excel, search it on a website and stored the collected data in same excel sheet.
institutionNames()
dataScreapping()
# print(appByDict1)
# print(affByDict1)
# print(appByDict2)
# print(affByDict2)
dataPosting()