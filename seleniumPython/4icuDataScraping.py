"""
The below program gets data from excel sheet and the look up for the data on a website and then scraps that data
and then insert it on a excel sheet.
"""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl as op
import time

drive = "" #! Please enter drive path.
fileName = "" #! Please enter file name.
path = drive + fileName

r = 3 #! Please change the row number to the row number from which you would like to start reading the file.
nameList = list()
def institutionsNames():
    """
    This function load all the data from a excel sheet for a particular column and stores it in a list named "nameList".
    Current design to get data from single column but can be chaged to get data from multiple column.
    """
    itn = op.load_workbook(path)
    itnSh = itn["Sheet1"]
    maxRow = itnSh.max_row
    for s in range(r, (maxRow + 10)):
        name = itnSh.cell(row = s, column = 3).value #! Please change the column number based on the column number in your file.
        if name != None:
            name = name.strip()
            nameList.append(name)
    itn.close()
    return nameList

accreditationsDict = dict()
def accreditationsScraping():
    """
    This function will scrap the data from the website and store it a dict name "accreditationsDict".
    It uses selenium webdriver for searching and getting the data.
    """
    chromeDriverPath = "" #! Please add your googlechrome driver path.
    chromeOptions = webdriver.ChromeOptions()
    #? Feel free to adjust the below argument based on your requirement.
    chromeOptions.add_argument("start-maximized")
    chromeOptions.add_argument("disable-infobars")
    chromeOptions.add_argument("headless")
    chromeOptions.add_argument("disable-popup-blocking")
    chromeOptions.add_experimental_option("detach",True)

    driver = webdriver.Chrome(chromeDriverPath, options=chromeOptions)
    driver.get("https://www.4icu.org/in/a-z/")

    #? Calling the above function here, so no need to call it seperately.
    institutionsNames()
    nameListLength = len(nameList)
    print(nameListLength)
    # curr = 0
    for i in nameList:
        # curr += 1
        # print(curr)
        driver.get("https://www.4icu.org/in/a-z/")

        try:
            schoolName = driver.find_element_by_link_text(i)
            schoolName.click()

            time.sleep(2)

            accreditations = driver.find_element_by_xpath("/html/body/div[3]/div[10]/div[2]/p[1]/a")
            accreditationsText = accreditations.text
            accreditationsDict[i] = accreditationsText
        except:
            pass
    return accreditationsDict

def dataPosting():
    """
    This functions open the same excel sheet from which we retrieve data
    and then will store the collected data in the new column.
    """
    itn = op.load_workbook(path)
    itnSh = itn["Sheet1"]
    maxRow = itnSh.max_row
    for s in range(r, (maxRow + 10)):
        name = itnSh.cell(row = s, column = 3).value #! Please change the column number based on the column number in your file.
        for key, val in accreditationsDict.items():
            if key == name:
                accreditationsValue = itnSh.cell(row = s, column = 10) #! Please change the column number based on the column number in your file.
                accreditationsValue.value = val
    itn.save(path)
    itn.close()

#? Calling the above functions, which will get data from excel, search it on a website and stored the collected data in same excel sheet.
accreditationsScraping()
dataPosting()
