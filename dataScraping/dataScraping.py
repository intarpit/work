"""
The below program gets data from excel sheet and the look up for the data on a website and then scraps that data
and then insert it on a excel sheet.
! ? Currently work in progress.
"""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl as op
import time

drive = "" #! Please enter drive path.
fileName = "" #! Please enter file name.
path = drive + fileName

institutionNameFile = "InstitutionsNames"
instNamePath = drive + institutionNameFile

st = 2
nameDict = dict()

def InstitutionsNames():
    itn = op.load_workbook(institutionNameFile)
    itnSh = itn.get_sheet_by_name("Sheet1")
    maxRow = itn.max_row
    for s in range(2, (maxRow + 10)):
        name = itnSh.cell(row = s, column = 2).value
        if name != None:
            name = name.strip()
            nameDict.append(name)
    return nameDict


chromeOptions = webdriver.ChromeOptions()
chromeOptions.add_argument("start-maximized")
chromeOptions.add_argument("disable-infobars")
chromeOptions.add_argument("disable-popup-blocking")
chromeOptions.add_experimental_option("detach",True)

driver = webdriver.Chrome(chrome_options=chromeOptions)
driver.get("https://www.edufever.com/")

for i in nameDict:
    seachSign = driver.find_element_by_xpath("//*[@id='menu-mymenu-1']/li[9]/a")
    seachSign.click()

    searchBar = driver.find_element_by_id("q")
    searchBar.send_keys(i)


time.sleep(500)
