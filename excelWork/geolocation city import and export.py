
"""
The below program get/input data from a excel sheet and will write/print the unique data on text document.
"""
import openpyxl as op
import os

drive = "" #! Please enter the drive path.
geo_input_file = "" #! Please enter the input file name.
geo_output_file = "" #! Please enter the output file name.

geo_input_path = drive + geo_input_file
geo_output_path = drive + geo_output_file


# * Loading all the district from geolocation excelsheet.

geolocation_city = []
wb1 = op.load_workbook(geo_input_path)
ws1 = wb1.get_sheet_by_name("Sheet1")
rm1 = ws1.max_row
for loc in range(2, (rm1 + 1)):
    geo_cit = ws1.cell(row = loc, column = 1).value
    if geo_cit not in geolocation_city:
        geolocation_city.append(geo_cit)
        output = open(geo_output_path, 'a')
        output.write(geo_cit + "\n")
        output.close()
print("All Done")
