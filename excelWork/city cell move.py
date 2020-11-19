"""
The below program capitalize all the data and will move to different column if the requirments are satisfied.
"""
import openpyxl as op
import string

drive = "" #! Please enter the drive address
filename = "" #! Please enter the filename.
path = drive + filename

wb = op.load_workbook(path)
ws = wb.get_sheet_by_name("Sheet1")
rm = ws.max_row

for s in range(2, (rm + 1)):
    old_city = ws.cell(row = s, column = 10).value
    if old_city != None:
        if len(old_city) < 15:
            cit = string.capwords(old_city)
            new_city = ws.cell(row = s, column = 11)
            new_city.value = cit
            extra_data = ws.cell(row = s, column = 10)
            extra_data.value = None
wb.save(path)
print("All Done")

