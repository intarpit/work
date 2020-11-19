"""
The below program input the data from a text file, matches the data on a excel sheet and will return the values, 
where their is a mismatch of information.
"""
import openpyxl as op

filename = "" #! Please enter file name.
drive = "" #! Please enter drive path.
geo_file = "" #! Please enter file name.
wrong_district = "" #! Please enter the output file name.


geo_path = drive + geo_file
path = drive + filename
district_path = drive + wrong_district

# * gettting district data from text file into a list
geo_district = []
with open (geo_path, 'r') as gp:
    geo_district = gp.read().splitlines()

# * loading the excel sheet and validating if the name is the excel sheet is correct or not.
wb1 = op.load_workbook(path)
ws1 = wb1.get_sheet_by_name("Sheet1")
rm1 = ws1.max_row

wb2 = op.load_workbook(district_path)
ws2 = wb2.get_sheet_by_name("Sheet1")
t = 2

non_existing_district = []
for s in range(2, (rm1 + 1)):
    district = ws1.cell(row = s, column = 12).value
    state = ws1.cell(row = s, column = 13).value
    if district != None:
        if district not in geo_district:
            if district not in non_existing_district:
                non_existing_district.append(district)
                district2 = ws2.cell(row = t, column = 1)
                district2.value = district
                state2 = ws2.cell(row = t, column = 2)
                state2.value = state
                t += 1
wb2.save(district_path)            
print("All Done")

