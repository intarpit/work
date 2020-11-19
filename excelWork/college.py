import openpyxl as op

""" 
The below file checks for duplicate data in a column and will return all the values, where duplicate data was found.
"""

# ! Please change the file name and drive name, based on your drive location for input excel sheet.
# ! Please change the text file name and drive name (if text file is in different folder or drive), based on your drive location for out text file.
# ! When the program is run and it return no name in the console or in the text file, that means the file is clean no more duplicates.

# ? Input file name
filename = "" #!Please enter the file name.

# ? Drive name
drive = "" #! Please enter the drive name.
path = drive + filename

# * Opening and loading the workbook and worksheet. Please make sure that worksheet name is same as the name in the second line.
workbook = op.load_workbook(path)
worksheet = workbook.get_sheet_by_name("Sheet1")
max_rows = worksheet.max_row

# ? text file name
text_file = "duplicate_names.txt"

# * This list will store the name and will stop the program from entering same name multiple times in the text file and the console.
duplicate_names = []

# * Running a loop to read all the names on excel sheet one by one.
for r in range(2, (max_rows + 1)):
    # * Getting the value of organization name, district and city.
    names1 = worksheet.cell(row = r, column = 2).value
    district1 = worksheet.cell(row = r, column = 11).value
    city1 = worksheet.cell(row = r, column = 10).value
    # * Converting all the above value to lowercase and striping all the white spaces from front and behind.
    names1 = names1.lower().strip()
    # * Ignoring the the rows if the values are blank.
    if district1 != None:
        district1 = district1.lower().strip()
    if city1 != None:    
        city1 = city1.lower().strip()
    # * Running a loop to read all the names, starting from next row of the above data.
    for s in range((r+1), (max_rows +1)):
        # * Getting the value of orgainzation name, district and city in the next row.
        names2 = worksheet.cell(row = s, column = 2).value
        district2 = worksheet.cell(row = s, column = 11).value
        city2 = worksheet.cell(row = s, column = 10).value
        # * Converting all the above value to lowercase and stripping all the white spaces from front and behind. 
        names2 = names2.lower().strip()
        # * Ignoring the values if the values are blank.
        if district2 != None:
            district2 = district2.lower().strip()
        if city2 != None:
            city2 = city2.lower().strip()
        # * Comparing the each and every names on the sheet.
        if names1 == names2:
            # * Validating if district is null, if yes then ignoring it.
            if district1 != None:
            # * Comparing the district, if the names of the organization are same.
                if (district1 == district2):
                # * Validating that the names already doesn't exist in duplicate_name list.
                # * So that we don't have multiple entries in our text file and console. 
                    if names1 not in duplicate_names:
                        # * Opening the text file, all the paramaters are global variable, which are mentioned before the start of any loop.
                        dm = open(drive + text_file, 'a')
                        # * Adding the name of the organization in the duplicate_name list.
                        duplicate_names.append(names1)
                        # * Printing out all the names, which have same organization name and also have same city or district.
                        print(names1)
                        # * Writting the above names in our text file. So that it can be shared.
                        dm.write(names1 + "\n")
                        # * Closing the text file.
                        dm.close()
            # * Validating that the value of city is null, if yes than ignoring it.
            elif city1 != None:
                # * Comparing the city, if the names of the organization are same.
                if (city1 == city2):
                # * Validating that the names already doesn't exist in duplicate_name list.
                # * So that we don't have multiple entries in our text file and console. 
                    if names1 not in duplicate_names:
                        # * Opening the text file, all the paramaters are global variable, which are mentioned before the start of any loop.
                        dm = open(drive + text_file, 'a')
                        # * Adding the name of the organization in the duplicate_name list.
                        duplicate_names.append(names1)
                        # * Printing out all the names, which have same organization name and also have same city or district.
                        print(names1)
                        # * Writting the above names in our text file. So that it can be shared.
                        dm.write(names1 + "\n")
                        # * Closing the text file.
                        dm.close()
            else:
                pass
# * Printing out the duplicate_name list but its not necessary can be commented out.
print(duplicate_names)